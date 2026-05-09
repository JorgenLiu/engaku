"""formula_graph.py -- Formula dependency graph builder for xlsx-analyze skill.

Parses formulas with openpyxl.formula.Tokenizer and builds a cell-to-cell
dependency graph without evaluating any formula.

Usage:
    python formula_graph.py <path> [--sheet NAME] [--focus CELL]
                            [--direction dependencies|dependents|both]
                            [--max-depth N] [--format json|markdown]

Supported file types: .xlsx, .xlsm
"""
import argparse
import json
import os
import re
import sys

_INSTALL_CMD = (
    "python -m pip install -r .github/skills/xlsx-analyze/requirements-py38.txt"
)

# Max cells to expand from a range reference before treating it as a range node
_MAX_RANGE_EXPAND = 500

# Pattern: optional sheet prefix (Sheet1! or 'My Sheet'!) + cell/range reference
_SHEET_PREFIX_RE = re.compile(
    r"^(?:\[.*?\])?(?:'[^']+'|[A-Za-z0-9_.]+)!"
)
_CELL_RANGE_RE = re.compile(
    r"^\$?[A-Z]{1,3}\$?[0-9]+(:\$?[A-Z]{1,3}\$?[0-9]+)?$"
)


def _check_deps():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.stderr.write(
            "Error: openpyxl is not installed.\n"
            "Install with: {}\n".format(_INSTALL_CMD)
        )
        sys.exit(1)


def _split_sheet_ref(token_value, default_sheet):
    """Split 'Sheet!Ref' or '[file]Sheet!Ref' into (sheet, ref, is_external)."""
    m = _SHEET_PREFIX_RE.match(token_value)
    if not m:
        return default_sheet, token_value, False

    prefix = m.group(0)
    ref = token_value[len(prefix):]
    is_external = prefix.startswith("[")

    # Extract sheet name: strip leading [file] if present
    raw = prefix.rstrip("!")
    if raw.startswith("["):
        bracket_end = raw.index("]")
        raw = raw[bracket_end + 1:]
    sheet = raw.strip("'")
    return sheet, ref, is_external


def _col_to_num(col_str):
    col_str = col_str.lstrip("$").upper()
    n = 0
    for ch in col_str:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _num_to_col(n):
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(rem + ord("A")) + s
    return s


def _expand_range(ref, sheet, max_cells):
    """Expand 'A1:C3' to individual cell strings. Returns (cells, truncated)."""
    ref_clean = ref.replace("$", "")
    if ":" not in ref_clean:
        return [ref_clean], False

    start, end = ref_clean.split(":", 1)
    m1 = re.match(r"([A-Z]+)([0-9]+)", start.upper())
    m2 = re.match(r"([A-Z]+)([0-9]+)", end.upper())
    if not m1 or not m2:
        return [ref], False

    c1, r1 = m1.groups()
    c2, r2 = m2.groups()
    c1n, c2n = _col_to_num(c1), _col_to_num(c2)
    r1n, r2n = int(r1), int(r2)

    total = (c2n - c1n + 1) * (r2n - r1n + 1)
    if total > max_cells:
        return [ref], True

    cells = []
    for r in range(r1n, r2n + 1):
        for cn in range(c1n, c2n + 1):
            cells.append("{}{}".format(_num_to_col(cn), r))
    return cells, False


def _parse_formula_refs(formula, sheet_name):
    """Return (refs, warnings) where refs is a list of dicts."""
    from openpyxl.formula import tokenize

    try:
        tok = tokenize.Tokenizer(formula)
    except Exception as exc:
        return [], ["Tokenizer error: {}".format(exc)]

    refs = []
    warnings = []

    for token in tok.items:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue

        value = token.value.strip()
        sheet, ref, is_external = _split_sheet_ref(value, sheet_name)

        if not _CELL_RANGE_RE.match(ref.replace("$", "")):
            # Might be a named range or unsupported construct
            warnings.append("Unrecognized reference token: {}".format(value))
            continue

        refs.append({
            "target": value,
            "ref": ref,
            "sheet": sheet,
            "token_type": token.type,
            "unresolved": is_external or (sheet != sheet_name),
        })

    return refs, warnings


def _build_graph(path, sheet_name):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)

    if sheet_name not in wb.sheetnames:
        sys.stderr.write(
            "Error: sheet '{}' not found. Available sheets: {}\n".format(
                sheet_name, ", ".join(wb.sheetnames)
            )
        )
        wb.close()
        sys.exit(1)

    ws = wb[sheet_name]
    nodes = set()
    edges = []
    all_warnings = []

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue

            source = "{}!{}".format(sheet_name, cell.coordinate)
            nodes.add(source)

            refs, warns = _parse_formula_refs(cell.value, sheet_name)
            all_warnings.extend(warns)

            for ref_info in refs:
                target_node = "{}!{}".format(ref_info["sheet"], ref_info["ref"])
                nodes.add(target_node)
                edges.append({
                    "source": source,
                    "target": ref_info["ref"],
                    "sheet": ref_info["sheet"],
                    "formula": cell.value,
                    "token_type": ref_info["token_type"],
                    "unresolved": ref_info["unresolved"],
                })

    wb.close()
    return sorted(nodes), edges, all_warnings


def _filter_by_focus(nodes, edges, focus, direction, sheet_name):
    focus_full = "{}!{}".format(sheet_name, focus.upper().replace("$", ""))
    filtered = []
    for edge in edges:
        src_matches = edge["source"] == focus_full
        tgt_full = "{}!{}".format(edge["sheet"], edge["target"].replace("$", ""))
        tgt_matches = tgt_full == focus_full

        if direction in ("dependencies", "both") and src_matches:
            filtered.append(edge)
        elif direction in ("dependents", "both") and tgt_matches:
            filtered.append(edge)

    used = set()
    for e in filtered:
        used.add(e["source"])
        used.add("{}!{}".format(e["sheet"], e["target"]))
    return sorted(used), filtered


def _to_markdown(data):
    lines = [
        "# Formula Graph: {}".format(data.get("file", "")),
        "**Sheet:** {} | **Nodes:** {} | **Edges:** {}".format(
            data.get("sheet", ""),
            len(data.get("nodes", [])),
            len(data.get("edges", [])),
        ),
    ]
    if data.get("focus"):
        lines.append("**Focus:** {} ({})".format(data["focus"], data.get("direction", "")))
    if data.get("warnings"):
        lines.append("")
        lines.append("## Warnings")
        for w in data["warnings"]:
            lines.append("- {}".format(w))
    lines.append("")
    lines.append("## Edges")
    lines.append("")
    lines.append("| Source | Target | Sheet | Unresolved |")
    lines.append("|--------|--------|-------|------------|")
    for e in data.get("edges", []):
        lines.append("| {} | {} | {} | {} |".format(
            e["source"], e["target"], e["sheet"], e["unresolved"]
        ))
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Build a formula dependency graph from an Excel workbook."
    )
    parser.add_argument("path", help="Path to .xlsx or .xlsm file")
    parser.add_argument(
        "--sheet", default=None,
        help="Sheet name (default: first sheet)"
    )
    parser.add_argument(
        "--focus", default=None,
        help="Focus on a specific cell (e.g. B5)"
    )
    parser.add_argument(
        "--direction", choices=["dependencies", "dependents", "both"],
        default="both", help="Graph direction when --focus is set (default: both)"
    )
    parser.add_argument(
        "--max-depth", type=int, default=5,
        help="Max traversal depth (default: 5; reserved for future use)"
    )
    parser.add_argument(
        "--format", choices=["json", "markdown"], default="json",
        dest="fmt", help="Output format (default: json)"
    )
    args = parser.parse_args()

    if not os.path.exists(args.path):
        sys.stderr.write("Error: file not found: {}\n".format(args.path))
        sys.exit(1)

    ext = os.path.splitext(args.path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        sys.stderr.write(
            "Error: formula_graph.py only supports .xlsx and .xlsm files.\n"
        )
        sys.exit(1)

    _check_deps()

    import openpyxl
    wb_check = openpyxl.load_workbook(args.path, read_only=True, data_only=False)
    sheet_name = args.sheet if args.sheet else wb_check.sheetnames[0]
    wb_check.close()

    nodes, edges, warnings = _build_graph(args.path, sheet_name)

    if args.focus:
        nodes, edges = _filter_by_focus(nodes, edges, args.focus, args.direction, sheet_name)

    data = {
        "file": os.path.basename(args.path),
        "sheet": sheet_name,
        "focus": args.focus,
        "direction": args.direction,
        "nodes": list(nodes),
        "edges": edges,
        "warnings": warnings,
    }

    if args.fmt == "markdown":
        print(_to_markdown(data))
    else:
        print(json.dumps(data, indent=2))


if __name__ == "__main__":
    main()
