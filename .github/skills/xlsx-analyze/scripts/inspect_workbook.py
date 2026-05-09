"""inspect_workbook.py -- Workbook structure inspector for xlsx-analyze skill.

Usage:
    python inspect_workbook.py <path> [--format json|markdown]
                               [--max-sheets N] [--max-formulas N]

Supported file types: .xlsx, .xlsm, .csv, .tsv
"""
import argparse
import json
import os
import sys

_INSTALL_CMD = (
    "python -m pip install -r .github/skills/xlsx-analyze/requirements-py38.txt"
)


def _check_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.stderr.write(
            "Error: openpyxl is not installed.\n"
            "Install with: {}\n".format(_INSTALL_CMD)
        )
        sys.exit(1)


def _inspect_xlsx(path, max_sheets, max_formulas):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    sheet_names = wb.sheetnames
    sheets = []
    sample_formulas = []

    for sheet_name in sheet_names[:max_sheets]:
        ws = wb[sheet_name]
        formula_count = 0

        try:
            merged = len(list(ws.merged_cells.ranges))
        except Exception:
            merged = 0

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    if len(sample_formulas) < max_formulas:
                        sample_formulas.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                        })

        sheets.append({
            "name": sheet_name,
            "min_row": ws.min_row or 0,
            "max_row": ws.max_row or 0,
            "min_col": ws.min_column or 0,
            "max_col": ws.max_column or 0,
            "merged_cells": merged,
            "formula_count": formula_count,
        })

    wb.close()
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    return {
        "file": os.path.basename(path),
        "format": ext,
        "sheet_count": len(sheet_names),
        "sheets": sheets,
        "sample_formulas": sample_formulas,
    }


def _inspect_csv(path):
    import csv

    ext = os.path.splitext(path)[1].lower()
    delimiter = "\t" if ext == ".tsv" else ","
    file_size = os.path.getsize(path)

    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    headers = rows[0] if rows else []
    row_count = len(rows) - 1 if rows else 0

    return {
        "file": os.path.basename(path),
        "format": ext.lstrip("."),
        "delimiter": repr(delimiter),
        "header_count": len(headers),
        "headers": headers,
        "row_count": row_count,
        "file_size_bytes": file_size,
    }


def _to_markdown(data):
    lines = ["# Workbook Inspection: {}".format(data.get("file", "")), ""]
    lines.append("**Format:** {}".format(data.get("format", "")))

    if "sheets" in data:
        lines.append("**Sheet count:** {}".format(data.get("sheet_count", "")))
        lines.append("")
        lines.append("## Sheets")
        lines.append("")
        lines.append("| Sheet | Rows | Cols | Merged | Formulas |")
        lines.append("|-------|------|------|--------|----------|")
        for s in data["sheets"]:
            lines.append("| {} | {}-{} | {}-{} | {} | {} |".format(
                s["name"],
                s["min_row"], s["max_row"],
                s["min_col"], s["max_col"],
                s["merged_cells"],
                s["formula_count"],
            ))
        if data.get("sample_formulas"):
            lines.append("")
            lines.append("## Sample Formulas")
            for sf in data["sample_formulas"]:
                lines.append("- **{}!{}**: `{}`".format(
                    sf["sheet"], sf["cell"], sf["formula"]
                ))
    else:
        lines.append("**Delimiter:** {}".format(data.get("delimiter", "")))
        lines.append("**Headers ({}):** {}".format(
            data.get("header_count", 0),
            ", ".join(data.get("headers", [])),
        ))
        lines.append("**Row count:** {}".format(data.get("row_count", 0)))
        lines.append("**File size:** {} bytes".format(data.get("file_size_bytes", 0)))

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Inspect an Excel workbook or delimited file."
    )
    parser.add_argument("path", help="Path to .xlsx, .xlsm, .csv, or .tsv file")
    parser.add_argument(
        "--format", choices=["json", "markdown"], default="json",
        dest="fmt", help="Output format (default: json)"
    )
    parser.add_argument(
        "--max-sheets", type=int, default=50,
        help="Max sheets to inspect (default: 50)"
    )
    parser.add_argument(
        "--max-formulas", type=int, default=10,
        help="Max sample formulas to collect (default: 10)"
    )
    args = parser.parse_args()

    if not os.path.exists(args.path):
        sys.stderr.write("Error: file not found: {}\n".format(args.path))
        sys.exit(1)

    ext = os.path.splitext(args.path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        _check_openpyxl()
        data = _inspect_xlsx(args.path, args.max_sheets, args.max_formulas)
    elif ext in (".csv", ".tsv"):
        data = _inspect_csv(args.path)
    else:
        sys.stderr.write(
            "Error: unsupported file type '{}'. "
            "Supported: .xlsx, .xlsm, .csv, .tsv\n".format(ext)
        )
        sys.exit(1)

    if args.fmt == "markdown":
        print(_to_markdown(data))
    else:
        print(json.dumps(data, indent=2))


if __name__ == "__main__":
    main()
